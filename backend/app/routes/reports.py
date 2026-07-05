from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from app.database.session import get_db
from app.models.models import Event, Registration, Attendance, Feedback, User
from app.core.security import get_current_user, RoleChecker

router = APIRouter(prefix="/api/reports", tags=["Reports Export"])

organizer_or_admin = RoleChecker(["organizer", "admin"])

@router.get("/attendance/{event_id}")
def export_attendance_report(
    event_id: int,
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(organizer_or_admin)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
        
    if current_user.role != "admin" and event.organizer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    registrations = db.query(Registration).filter(Registration.event_id == event_id).all()
    
    # Pre-process data
    data = []
    for r in registrations:
        attendance_record = db.query(Attendance).filter(Attendance.registration_id == r.id).first()
        checkin_time = attendance_record.check_in_time.strftime("%Y-%m-%d %H:%M:%S") if attendance_record else "Absent"
        data.append({
            "User ID": r.user.id,
            "Name": r.user.name,
            "Email": r.user.email,
            "Phone": r.user.phone or "N/A",
            "Organization": r.user.organization or "N/A",
            "Registration Date": r.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "Check-In Status": "Checked In" if r.checked_in else "Absent",
            "Check-In Time": checkin_time
        })
        
    filename = f"attendance_report_event_{event_id}_{datetime.date.today()}"
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys() if data else ["Status"])
        writer.writeheader()
        if data:
            writer.writerows(data)
        else:
            output.write("No participants registered yet.")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
        
    elif format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Style
        title_font = Font(name="Arial", size=16, bold=True, color="1B365D")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append([f"Attendance Report: {event.title}"])
        ws.merge_cells("A1:H1")
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = alignment
        ws.row_dimensions[1].height = 40
        
        headers = ["User ID", "Name", "Email", "Phone", "Organization", "Registration Date", "Status", "Check-in Time"]
        ws.append([]) # Empty row
        ws.append(headers)
        
        for col_idx in range(1, 9):
            ws.cell(row=3, column=col_idx).font = header_font
            ws.cell(row=3, column=col_idx).fill = header_fill
            ws.cell(row=3, column=col_idx).alignment = alignment
            
        for row in data:
            ws.append([
                row["User ID"],
                row["Name"],
                row["Email"],
                row["Phone"],
                row["Organization"],
                row["Registration Date"],
                row["Check-In Status"],
                row["Check-In Time"]
            ])
            
        # Resize columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = chr(64 + col[0].column) if col[0].column <= 26 else "A" # Quick limit
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
        
    elif format == "pdf":
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1B365D'),
            spaceAfter=15
        )
        body_style = styles['Normal']
        
        story = []
        story.append(Paragraph(f"Attendance Report: {event.title}", title_style))
        story.append(Paragraph(f"Venue: {event.venue} | Date: {event.start_date.strftime('%Y-%m-%d')}", body_style))
        story.append(Spacer(1, 15))
        
        # Build Table
        table_data = [["ID", "Name", "Email", "Status", "Check-In Time"]]
        for row in data:
            table_data.append([
                str(row["User ID"]),
                row["Name"][:15],
                row["Email"][:20],
                row["Check-In Status"],
                row["Check-In Time"]
            ])
            
        table = Table(table_data, colWidths=[30, 110, 150, 80, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F1F5F9')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        
        story.append(table)
        doc.build(story)
        pdf_buffer.seek(0)
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )

@router.get("/feedback/{event_id}")
def export_feedback_report(
    event_id: int,
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(organizer_or_admin)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
        
    if current_user.role != "admin" and event.organizer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Unauthorized")
        
    feedbacks = db.query(Feedback).filter(Feedback.event_id == event_id).all()
    
    data = []
    for f in feedbacks:
        data.append({
            "User Name": f.user.name,
            "Rating": f.rating,
            "Comment": f.comment or "No comment",
            "Sentiment": f.sentiment,
            "Date": f.created_at.strftime("%Y-%m-%d %H:%M:%S")
        })
        
    filename = f"feedback_report_event_{event_id}_{datetime.date.today()}"
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys() if data else ["Feedback Status"])
        writer.writeheader()
        if data:
            writer.writerows(data)
        else:
            output.write("No feedback submitted yet.")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
        
    elif format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append([f"Feedback Sentiment Analysis Report: {event.title}"])
        ws.append([])
        ws.append(["User Name", "Rating", "Comment", "Sentiment", "Submitted Date"])
        
        for row in data:
            ws.append([row["User Name"], row["Rating"], row["Comment"], row["Sentiment"], row["Date"]])
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
        
    elif format == "pdf":
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1B365D')
        )
        story = [
            Paragraph(f"Feedback Report: {event.title}", title_style),
            Paragraph(f"Generated on: {datetime.date.today()}", styles['Normal']),
            Spacer(1, 15)
        ]
        
        table_data = [["Name", "Rating", "Comment", "Sentiment"]]
        for row in data:
            table_data.append([
                row["User Name"],
                str(row["Rating"]),
                Paragraph(row["Comment"], styles['Normal']),
                row["Sentiment"]
            ])
            
        table = Table(table_data, colWidths=[100, 50, 270, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(table)
        doc.build(story)
        pdf_buffer.seek(0)
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )

@router.get("/analytics")
def export_analytics_report(
    format: str = Query("csv", pattern="^(csv|xlsx|pdf)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(organizer_or_admin)
):
    # Filter by organizer if not admin
    is_admin = current_user.role == "admin"
    org_id = current_user.id if not is_admin else None

    # Retrieve organizer's events or all events if admin
    events_q = db.query(Event)
    if org_id:
        events_q = events_q.filter(Event.organizer_id == org_id)
    events = events_q.all()

    # Pre-process data
    data = []
    for e in events:
        reg_count = db.query(Registration).filter(Registration.event_id == e.id).count()
        checked_in_count = db.query(Registration).filter(
            Registration.event_id == e.id,
            Registration.checked_in == True
        ).count()
        attendance_rate = float(round((checked_in_count / reg_count) * 100, 1)) if reg_count > 0 else 0.0
        
        # Feedback details
        feedbacks = db.query(Feedback).filter(Feedback.event_id == e.id).all()
        avg_rating = sum(f.rating for f in feedbacks) / len(feedbacks) if feedbacks else 0.0
        avg_rating = round(avg_rating, 1)

        data.append({
            "Event ID": e.id,
            "Title": e.title,
            "Category": e.category,
            "Venue": e.venue,
            "Capacity": e.capacity,
            "Registrations": reg_count,
            "Checked In": checked_in_count,
            "Attendance Rate (%)": attendance_rate,
            "Avg Rating (1-5)": avg_rating,
            "Start Date": e.start_date.strftime("%Y-%m-%d %H:%M:%S")
        })

    filename = f"analytics_report_{current_user.role}_{datetime.date.today()}"

    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys() if data else ["Status"])
        writer.writeheader()
        if data:
            writer.writerows(data)
        else:
            output.write("No events recorded yet.")
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )

    elif format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Event Analytics"
        
        # Style
        title_font = Font(name="Arial", size=16, bold=True, color="1B365D")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append([f"Event Analytics Report ({current_user.name})"])
        ws.merge_cells("A1:J1")
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = alignment
        ws.row_dimensions[1].height = 40
        
        headers = ["Event ID", "Title", "Category", "Venue", "Capacity", "Registrations", "Checked In", "Attendance Rate (%)", "Avg Rating (1-5)", "Start Date"]
        ws.append([]) # Empty row
        ws.append(headers)
        
        for col_idx in range(1, 11):
            ws.cell(row=3, column=col_idx).font = header_font
            ws.cell(row=3, column=col_idx).fill = header_fill
            ws.cell(row=3, column=col_idx).alignment = alignment
            
        for row in data:
            ws.append([
                row["Event ID"],
                row["Title"],
                row["Category"],
                row["Venue"],
                row["Capacity"],
                row["Registrations"],
                row["Checked In"],
                row["Attendance Rate (%)"],
                row["Avg Rating (1-5)"],
                row["Start Date"]
            ])
            
        # Resize columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = chr(64 + col[0].column) if col[0].column <= 26 else "A"
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )

    elif format == "pdf":
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1B365D')
        )
        story = [
            Paragraph(f"Event Analytics Report ({current_user.name})", title_style),
            Paragraph(f"Role: {current_user.role} | Generated on: {datetime.date.today()}", styles['Normal']),
            Spacer(1, 15)
        ]
        
        table_data = [["Title", "Category", "Regs", "Att Rate", "Rating", "Date"]]
        for row in data:
            table_data.append([
                row["Title"][:18],
                row["Category"],
                str(row["Registrations"]),
                f"{row['Attendance Rate (%)']}%",
                f"{row['Avg Rating (1-5)']}/5",
                row["Start Date"][:10]
            ])
            
        table = Table(table_data, colWidths=[150, 70, 50, 60, 50, 100])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(table)
        doc.build(story)
        pdf_buffer.seek(0)
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )
